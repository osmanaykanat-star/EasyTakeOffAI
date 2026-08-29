import openpyxl
import re
import datetime
import os
from typing import Optional, Dict, Any, List
from ..trades.trade_base import ProjectTakeoff, MaterialSpec, RoomTakeoff, TakeoffLineItem

class ExcelProposalParser:
    """
    Intelligent, universal parser for Excel Proposals and Takeoffs.
    Parses metadata, base bid rooms, line items, quantities, units, material specs, and exclusions.
    """
    @staticmethod
    def parse_excel(file_path: str) -> ProjectTakeoff:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        project = ProjectTakeoff(
            project_name=base_name,
            client_name="",
            client_company="",
            date_str=datetime.date.today().strftime("%m/%d/%Y"),
            material_specs={},
            rooms=[],
            exclusions=[]
        )
        
        current_floor = "MAIN FLOOR"
        current_room = None
        current_room_floor = "MAIN FLOOR"
        current_room_items = []
        
        in_rooms_section = True
        in_materials_section = False
        in_exclusions_section = False
        in_specs_table = False
        
        def save_current_room():
            nonlocal current_room, current_room_floor, current_room_items
            if current_room and current_room_items:
                project.rooms.append(RoomTakeoff(
                    room_name=current_room,
                    floor_name=current_room_floor,
                    length_ft=10.0,
                    width_ft=10.0,
                    ceiling_height_ft=9.0,
                    wall_tile_height_ft=0.0,
                    door_count=1,
                    items=list(current_room_items)
                ))
                current_room = None
                current_room_items = []

        prev_was_to = False
        col_count = max(ws.max_column or 0, 15)
        for r in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, col_count + 1)]
            row_str = " ".join([str(v) for v in row_vals if v is not None]).strip()
            if not row_str:
                continue
            
            # Check Date
            for v in row_vals:
                if isinstance(v, (datetime.datetime, datetime.date)):
                    project.date_str = v.strftime("%m/%d/%Y")
                elif isinstance(v, str) and re.match(r'^\d{4}-\d{2}-\d{2}', v.strip()):
                    try:
                        dt = datetime.datetime.strptime(v.strip()[:10], "%Y-%m-%d")
                        project.date_str = dt.strftime("%m/%d/%Y")
                    except Exception:
                        project.date_str = v.strip()
            
            # Check Header Info: To, Re
            row_has_to = False
            for c_idx, val in enumerate(row_vals):
                val_str = str(val or "").strip().lower()
                if val_str == "to:" or val_str.startswith("to:"):
                    for following in row_vals[c_idx + 1:]:
                        if following is not None and str(following).strip():
                            project.client_name = str(following).strip()
                            row_has_to = True
                            break
                elif val_str == "re:" or val_str.startswith("re:"):
                    for following in row_vals[c_idx + 1:]:
                        if following is not None and str(following).strip():
                            project.project_name = str(following).strip()
                            break
                elif prev_was_to and val_str and val_str not in ["to:", "re:", "dear", "proposal", "date"] and not project.client_company:
                    project.client_company = str(val).strip()
                    prev_was_to = False
                elif any(k in str(val or "").upper() for k in ["CORE FOUR", "SPK/LEWIS", "HITT CONTRACTING", "BERKS", "PRIME RENOVATIONS", "CROSS MANAGEMENT", "EVERGREEN", "CONSTRUCTION", "BUILDERS"]):
                    if not project.client_company and not any(k in str(val).upper() for k in ["SYMBOL", "UNIT PRICE", "BID AMOUNT"]):
                        project.client_company = str(val).strip()
            
            if row_has_to:
                prev_was_to = True
            elif prev_was_to and project.client_company:
                prev_was_to = False

            # Check Section Transitions & Termination of Room Table
            upper_row = row_str.upper()
            if upper_row.startswith("TOTAL") or upper_row.startswith("GRAND TOTAL") or "ABBREVIATIONS:" in upper_row or "EXCLUSIONS:" in upper_row or "MATERIALS INFORMATION:" in upper_row:
                save_current_room()
                in_rooms_section = False
                
            if "EXCLUSIONS:" in upper_row:
                in_exclusions_section = True
                in_materials_section = False
                in_specs_table = False
                continue
            elif "MATERIALS INFORMATION:" in upper_row or "MATERIAL INFORMATION" in upper_row:
                in_materials_section = True
                in_exclusions_section = False
                in_specs_table = False
                continue
            elif "CODE" in upper_row and ("MATERIAL SPEC" in upper_row or "SPECIFICATION" in upper_row):
                in_specs_table = True
                in_materials_section = False
                in_exclusions_section = False
                continue
            elif "BEST REGARDS" in upper_row:
                save_current_room()
                in_rooms_section = False
                in_exclusions_section = False
                in_materials_section = False
                in_specs_table = False
                continue
                
            if in_exclusions_section:
                if re.match(r'^\d+[\).]', row_str):
                    project.exclusions.append(row_str)
                continue
                
            # Spec table parser (e.g. CODE | MATERIAL SPEC | QUANTITY | UNIT | UNIT PRICE | NOTES)
            if in_specs_table:
                non_empty = [v for v in row_vals if v is not None and str(v).strip()]
                if len(non_empty) >= 2:
                    sym = str(non_empty[0]).strip()
                    if sym.upper() not in ["CODE", "TOTAL", "SYMBOL"]:
                        desc = str(non_empty[1]).strip()
                        unit = "SQ FT"
                        price = 0.0
                        notes = ""
                        for item_v in non_empty[2:]:
                            iv_str = str(item_v).strip().upper()
                            if iv_str in ["SQ FT", "SQFT", "SF", "LN FT", "LNFT", "LF", "PCS", "EA", "LS", "SETS"]:
                                unit = iv_str.replace("SQFT", "SQ FT").replace("LNFT", "LN FT")
                            elif isinstance(item_v, (int, float)):
                                if price == 0.0 and float(item_v) > 0 and float(item_v) < 1000:
                                    price = float(item_v)
                            elif isinstance(item_v, str) and item_v.strip().replace('$', '').replace(',', '.').replace('.', '', 1).isdigit():
                                try:
                                    p_val = float(item_v.strip().replace('$', '').replace(',', '.'))
                                    if price == 0.0: price = p_val
                                except Exception:
                                    pass
                            elif isinstance(item_v, str) and len(item_v.strip()) > 3:
                                notes += (" " + item_v.strip())
                                
                        project.material_specs[sym] = MaterialSpec(
                            symbol=sym,
                            description=desc,
                            unit=unit,
                            budget_price=price,
                            notes=notes.strip(),
                            trade="Tile & Stone"
                        )
                continue

            if in_materials_section:
                c1 = str(row_vals[0] or "").strip()
                if not c1:
                    for v in row_vals:
                        if v and str(v).strip():
                            c1 = str(v).strip()
                            break
                if c1 and c1.upper() not in ["MATERIALS INFORMATION:", "QUANTITY", "UNIT", "PRICE", "NOTES"]:
                    sym = c1
                    desc = c1
                    if " - " in c1:
                        parts = c1.split(" - ")
                        sym = parts[0].strip()
                        desc = " - ".join(parts[1:]).strip()
                    elif re.match(r'^\d+\.\s*([^:]+):', c1):
                        m = re.match(r'^\d+\.\s*([^:]+):\s*(.*)', c1)
                        sym = m.group(1).strip()
                        desc = m.group(2).strip()
                    elif "-" in c1:
                        parts = c1.split("-", 1)
                        sym = parts[0].strip()
                        desc = parts[1].strip()

                    unit = "SQ FT"
                    price = 0.0
                    notes = ""
                    for v in row_vals[1:]:
                        if v is not None:
                            v_str = str(v).strip().upper()
                            if v_str in ["SQ FT", "SQFT", "SF", "LN FT", "LNFT", "LF", "PCS", "EA"]:
                                unit = v_str.replace("SQFT", "SQ FT").replace("LNFT", "LN FT")
                            elif isinstance(v, (int, float)):
                                if price == 0.0 and float(v) > 0 and float(v) < 1000:
                                    price = float(v)
                            elif isinstance(v, str) and len(v.strip()) > 3:
                                notes += (" " + v.strip())
                                
                    if "TRIM" in sym.upper() or "BASE" in sym.upper():
                        unit = "PCS" if "PCS" in str(row_vals) else "LN FT"
                    elif "SADDLE" in sym.upper():
                        unit = "PCS"
                        
                    project.material_specs[sym] = MaterialSpec(
                        symbol=sym,
                        description=desc,
                        unit=unit,
                        budget_price=price,
                        notes=notes.strip(),
                        trade="Tile & Stone"
                    )
                continue
                
            if in_rooms_section:
                # Parse Floor header
                if len(row_vals) > 1 and row_vals[1] is not None:
                    val1 = str(row_vals[1]).strip().upper()
                    if ("FLOOR" in val1 or "LEVEL" in val1) and not any(row_vals[3:]):
                        save_current_room()
                        current_floor = val1
                        continue
                        
                # Parse Room header
                if len(row_vals) > 2 and row_vals[2] is not None:
                    val2 = str(row_vals[2]).strip().upper()
                    if val2 and (len(row_vals) <= 3 or row_vals[3] is None or str(row_vals[3]).strip() == ""):
                        save_current_room()
                        current_room = val2
                        current_room_floor = current_floor
                        continue
                        
                # Parse Line item row
                if len(row_vals) > 3 and row_vals[3] is not None:
                    sym = str(row_vals[3]).strip()
                    if sym and sym.upper() not in ["SYMBOL", "BASE BID", "TOTAL", "CODE"]:
                        finish_type = str(row_vals[4] or "").strip()
                        mat_type = str(row_vals[5] or "").strip()
                        work_type = "S&I"
                        qty = 0.0
                        mat_price = 0.0
                        labor_price = 0.0
                        
                        # Find work type
                        for v in row_vals[4:]:
                            if v is not None:
                                s = str(v).strip().upper()
                                if s in ["S&I", "IO"]:
                                    work_type = s
                                    break
                                    
                        # Find quantity (first float/int in columns 6 to 9)
                        for v in row_vals[6:10]:
                            if isinstance(v, (int, float)):
                                qty = float(v)
                                break
                            elif isinstance(v, str) and v.strip().replace(',', '').replace('.', '', 1).isdigit():
                                try:
                                    qty = float(v.strip().replace(',', ''))
                                    break
                                except Exception:
                                    pass

                        unit = "SQ FT"
                        for v in row_vals[7:10]:
                            if v is not None:
                                u_str = str(v).strip().upper()
                                if u_str in ["SQ FT", "SQFT", "SF"]:
                                    unit = "SQ FT"
                                    break
                                elif u_str in ["LN FT", "LNFT", "LF"]:
                                    unit = "LN FT"
                                    break
                                elif u_str in ["PCS", "PC", "EA"]:
                                    unit = "PCS"
                                    break
                            
                        item = TakeoffLineItem(
                            symbol=sym,
                            finish_type=finish_type,
                            material_type=mat_type if mat_type else "TILE",
                            work_type=work_type,
                            quantity=qty,
                            unit=unit,
                            material_price=mat_price,
                            labor_price=labor_price,
                            trade="Tile & Stone"
                        )
                        current_room_items.append(item)
                        
        save_current_room()
        
        if not project.exclusions:
            project.exclusions = [
                "1) Demolition / prep outside scope",
                "2) Epoxy Grout (unless specified)",
                "3) Air freight of any material"
            ]
            
        return project
