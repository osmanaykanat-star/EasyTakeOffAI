import os
import datetime
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..trades.trade_base import ProjectTakeoff

class ExcelProposalGenerator:
    """
    Generates pristine, easy-to-price Excel proposal files matching Osman's exact professional standard:
    - Exact column layout (A=Margin, B=Floor, C=Room, D=Symbol, E=Finish Type, F=Material Type, I=Work Type, J=Qty, K=Material Price, L=Labor Price, M=Bid Amount)
    - Dynamic Excel Formulas:
        * Line Item Bid Amount: =J*(K+L)
        * Room Subtotal: =SUM(M_start:M_end)
        * Base Bid (M20): Sum of all room subtotals
        * Summary Card (E16): =M20
        * Summary Totals: =SUMIF(D:D, Symbol, J:J)
    - Color-coded / distinct price input cells (K & L) with currency formatting ($#,##0.00)
    - Clean typography, proper auto-fit dimensions, abbreviations, exclusions, and material specs.
    """

    @staticmethod
    def generate_excel(project: ProjectTakeoff, output_path: str) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proposal"

        # Page Setup: Pristine Portrait formatting & Fit to 1 Page Wide
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2

        ws.print_title_rows = '19:19'

        # 1. Compact Column Dimensions (optimized for Portrait layout)
        col_widths = {
            'A': 3,
            'B': 11,
            'C': 16,
            'D': 10,
            'E': 20,
            'F': 16,
            'G': 1,
            'H': 1,
            'I': 8,
            'J': 10,
            'K': 11,
            'L': 11,
            'M': 13
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # 2. Typography & Color Palette
        font_title = Font(name="Calibri", size=14, bold=True, color="1F2937")
        font_header = Font(name="Calibri", size=11, bold=True, color="1E293B")
        font_floor = Font(name="Calibri", size=11, bold=True, color="0F172A")
        font_room = Font(name="Calibri", size=10, bold=True, color="1E293B")
        font_regular = Font(name="Calibri", size=10, color="334155")
        font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
        font_sub = Font(name="Calibri", size=9, italic=True, color="64748B")

        # Fills
        fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Classic Soft Blue-Grey
        fill_floor = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # Light slate
        fill_room = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")   # Very soft slate
        fill_price_input = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid") # Soft warm amber/yellow for pricing input
        fill_bid = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")    # Soft green for calculated bid amount
        fill_base_bid = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Highlight gold for Base Bid

        # Borders
        border_thin_gray = Side(border_style="thin", color="CBD5E1")
        border_box = Border(left=border_thin_gray, right=border_thin_gray, top=border_thin_gray, bottom=border_thin_gray)
        border_top_thin = Border(top=Side(border_style="thin", color="1E293B"))
        border_double_bottom = Border(bottom=Side(border_style="double", color="0F172A"), top=Side(border_style="thin", color="0F172A"))

        # Alignments
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 3. Header Block
        ws['A1'] = "PROPOSAL"
        ws['A1'].font = font_title

        date_val = project.date_str if project.date_str else datetime.date.today().strftime("%m/%d/%Y")
        ws['B2'] = date_val
        ws['B2'].font = font_regular

        to_company = (project.client_company or "").strip()
        attn_person = (project.client_name or "").strip()

        if not to_company or to_company.upper() in ["GENERAL CONTRACTOR", "COMMERCIAL CONSTRUCTION", "CLIENT COMPANY", ""]:
            if any(k in attn_person.upper() for k in ["LLC", "INC", "CORP", "CONSTRUCTION", "BUILDERS", "GROUP", "MANAGEMENT", "PARTNERS", "DEVELOPMENT", "HOLDINGS"]):
                to_company = attn_person
                attn_person = "Project Estimator / Manager"
            else:
                to_company = "General Contractor / Construction Manager"

        if not attn_person:
            attn_person = "Project Manager"

        ws['A4'] = "To:"
        ws['A4'].font = font_bold
        ws['B4'] = to_company
        ws['B4'].font = font_bold

        ws['A5'] = "Attn:"
        ws['A5'].font = font_bold
        ws['B5'] = attn_person
        ws['B5'].font = font_bold

        ws['A7'] = "Re:"
        ws['A7'].font = font_bold
        ws['B7'] = project.project_name
        ws['B7'].font = font_bold

        client_first = attn_person.split()[0].title() if attn_person and attn_person not in ["Project Manager", "Project Estimator / Manager"] else "Sir/Madam"
        ws['A11'] = f"Dear {client_first},"
        ws['A11'].font = font_regular

        ws['A13'] = "We hereby propose to supply and install STONE & TILE work at the above location as per your plans and specs for the sum of $............................."
        ws['A13'].font = font_regular

        ws['A15'] = "Summary:"
        ws['A15'].font = font_bold

        ws['A16'] = "Base Bid"
        ws['A16'].font = font_bold
        ws['E16'] = "=M20"
        ws['E16'].font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
        ws['E16'].alignment = align_left
        ws['E16'].number_format = '$#,##0.00'

        # 4. Table Headers (Row 19)
        headers = [
            ("B19", "Floor"),
            ("C19", "Room"),
            ("D19", "Symbol"),
            ("E19", "Finish Type"),
            ("F19", "Material Type"),
            ("I19", "Work Type"),
            ("J19", "Material Quantity"),
            ("K19", "Material Unit Price"),
            ("L19", "Labor Unit Price"),
            ("M19", "Bid Amount")
        ]
        for cell_ref, text in headers:
            cell = ws[cell_ref]
            cell.value = text
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = border_box
        ws.row_dimensions[19].height = 28

        # 5. Base Bid Row (Row 20)
        ws['A20'] = "BASE BID"
        ws['A20'].font = font_title
        ws['M20'].font = font_title
        ws['M20'].alignment = align_right
        ws['M20'].number_format = '$#,##0.00'
        ws['M20'].fill = fill_base_bid
        ws['M20'].border = border_double_bottom

        current_row = 21
        room_subtotal_rows: List[int] = []
        item_rows_map: Dict[str, List[int]] = {} # symbol -> list of row indices for formula linking

        # Group rooms by floor
        floors_dict: Dict[str, list] = {}
        for room in project.rooms:
            floors_dict.setdefault(room.floor_name, []).append(room)

        first_data_row = 21

        for floor_name, floor_rooms in floors_dict.items():
            # Floor Row Header
            f_cell = ws.cell(row=current_row, column=2, value=floor_name.upper())
            f_cell.font = font_floor
            for c in range(2, 14):
                ws.cell(row=current_row, column=c).fill = fill_floor
                ws.cell(row=current_row, column=c).border = border_box
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            for room in floor_rooms:
                room_header_row = current_row
                r_cell = ws.cell(row=room_header_row, column=3, value=room.room_name.upper())
                r_cell.font = font_room
                for c in range(3, 14):
                    ws.cell(row=room_header_row, column=c).fill = fill_room
                    ws.cell(row=room_header_row, column=c).border = border_box
                ws.row_dimensions[room_header_row].height = 20
                current_row += 1

                room_items_start = current_row
                for item in room.items:
                    row_idx = current_row
                    item_rows_map.setdefault(item.symbol, []).append(row_idx)

                    # B: Floor Name
                    b_cell = ws.cell(row=row_idx, column=2, value=floor_name)
                    b_cell.font = font_regular
                    b_cell.alignment = align_left
                    b_cell.border = border_box

                    # C: Room Name
                    c_cell = ws.cell(row=row_idx, column=3, value=room.room_name)
                    c_cell.font = font_regular
                    c_cell.alignment = align_left
                    c_cell.border = border_box

                    # D: Symbol
                    d_cell = ws.cell(row=row_idx, column=4, value=item.symbol)
                    d_cell.font = font_bold
                    d_cell.alignment = align_left
                    d_cell.border = border_box

                    # E: Finish Type / Location Description
                    e_cell = ws.cell(row=row_idx, column=5, value=item.finish_type)
                    e_cell.font = font_regular
                    e_cell.alignment = align_left
                    e_cell.border = border_box

                    # F: Material Type
                    f_cell = ws.cell(row=row_idx, column=6, value=item.material_type)
                    f_cell.font = font_regular
                    f_cell.alignment = align_left
                    f_cell.border = border_box

                    # I: Work Type
                    i_cell = ws.cell(row=row_idx, column=9, value=item.work_type)
                    i_cell.font = font_regular
                    i_cell.alignment = align_center

                    # J: Material Quantity
                    j_cell = ws.cell(row=row_idx, column=10, value=item.quantity)
                    j_cell.font = font_bold
                    j_cell.alignment = align_right
                    j_cell.number_format = '#,##0.00'

                    # K: Material Unit Price (User Input Field)
                    k_cell = ws.cell(row=row_idx, column=11, value=item.material_price)
                    k_cell.font = font_regular
                    k_cell.alignment = align_right
                    k_cell.number_format = '$#,##0.00'
                    k_cell.fill = fill_price_input

                    # L: Labor Unit Price (User Input Field)
                    l_cell = ws.cell(row=row_idx, column=12, value=item.labor_price)
                    l_cell.font = font_regular
                    l_cell.alignment = align_right
                    l_cell.number_format = '$#,##0.00'
                    l_cell.fill = fill_price_input

                    # M: Bid Amount Dynamic Formula: =J*(K+L)
                    m_cell = ws.cell(row=row_idx, column=13, value=f"=J{row_idx}*(K{row_idx}+L{row_idx})")
                    m_cell.font = font_regular
                    m_cell.alignment = align_right
                    m_cell.number_format = '$#,##0.00'
                    m_cell.fill = fill_bid

                    # Set borders
                    for col_idx in [4, 5, 6, 9, 10, 11, 12, 13]:
                        ws.cell(row=row_idx, column=col_idx).border = border_box

                    ws.row_dimensions[row_idx].height = 19
                    current_row += 1

                room_items_end = current_row - 1
                if room_items_end >= room_items_start:
                    # Set room subtotal formula at column M on the room header row
                    room_subtotal_cell = ws.cell(row=room_header_row, column=13, value=f"=SUM(M{room_items_start}:M{room_items_end})")
                    room_subtotal_cell.font = font_bold
                    room_subtotal_cell.alignment = align_right
                    room_subtotal_cell.number_format = '$#,##0.00'
                    room_subtotal_cell.border = border_box
                    room_subtotal_rows.append(room_header_row)

        last_data_row = current_row - 1

        # Connect Base Bid formula in M20
        if room_subtotal_rows:
            ws['M20'] = "=" + "+".join([f"M{r}" for r in room_subtotal_rows])
        else:
            ws['M20'] = 0.0

        current_row += 2

        # 6. Total Quantities by Symbol Table (Dynamic SUMIF Formulas)
        ws.cell(row=current_row, column=1, value="TOTAL QUANTITIES BY MATERIAL").font = font_title
        current_row += 1

        total_header_row = current_row
        ws.cell(row=total_header_row, column=4, value="Symbol").font = font_header
        ws.cell(row=total_header_row, column=4).fill = fill_header
        ws.cell(row=total_header_row, column=4).alignment = align_left
        ws.cell(row=total_header_row, column=4).border = border_box

        ws.cell(row=total_header_row, column=5, value="Description / Specification").font = font_header
        ws.cell(row=total_header_row, column=5).fill = fill_header
        ws.cell(row=total_header_row, column=5).alignment = align_left
        ws.cell(row=total_header_row, column=5).border = border_box

        ws.cell(row=total_header_row, column=10, value="Total Quantity").font = font_header
        ws.cell(row=total_header_row, column=10).fill = fill_header
        ws.cell(row=total_header_row, column=10).alignment = align_right
        ws.cell(row=total_header_row, column=10).border = border_box

        ws.cell(row=total_header_row, column=11, value="Unit").font = font_header
        ws.cell(row=total_header_row, column=11).fill = fill_header
        ws.cell(row=total_header_row, column=11).alignment = align_center
        ws.cell(row=total_header_row, column=11).border = border_box
        current_row += 1

        # Aggregate total by symbol for list
        sym_totals: Dict[str, Dict[str, Any]] = {}
        for room in project.rooms:
            for item in room.items:
                if item.symbol not in sym_totals:
                    sym_totals[item.symbol] = {"qty": 0.0, "unit": item.unit}
                sym_totals[item.symbol]["qty"] += item.quantity

        for sym, data in sorted(sym_totals.items()):
            row_idx = current_row
            # Symbol
            s_c = ws.cell(row=row_idx, column=4, value=sym)
            s_c.font = font_bold
            s_c.border = border_box

            # Description
            spec = project.material_specs.get(sym)
            desc = spec.description if spec and spec.description else "Standard Specification"
            d_c = ws.cell(row=row_idx, column=5, value=desc)
            d_c.font = font_regular
            d_c.border = border_box

            # Formula for Total Quantity: =SUMIF(D$21:D$last, "sym", J$21:J$last)
            q_c = ws.cell(row=row_idx, column=10, value=f'=SUMIF(D${first_data_row}:D${last_data_row}, "{sym}", J${first_data_row}:J${last_data_row})')
            q_c.font = font_bold
            q_c.alignment = align_right
            q_c.number_format = '#,##0.00'
            q_c.border = border_box

            # Unit
            u_c = ws.cell(row=row_idx, column=11, value=data["unit"])
            u_c.font = font_regular
            u_c.alignment = align_center
            u_c.border = border_box

            current_row += 1

        current_row += 2

        # 7. Abbreviations
        ws.cell(row=current_row, column=1, value="Abbreviations:").font = font_bold
        current_row += 1
        ws.cell(row=current_row, column=1, value="   S&I: Supply & Install").font = font_regular
        current_row += 1
        ws.cell(row=current_row, column=1, value="   IO: Install Only").font = font_regular
        current_row += 2

        # 8. Exclusions
        ws.cell(row=current_row, column=1, value="Exclusions:").font = font_bold
        current_row += 1
        exclusions = project.exclusions or [
            "1) Air freight any material.",
            "2) Premium/Overtime labor unless agreed in writing.",
            "3) Structural subfloor repair or major crack isolation beyond standard prep."
        ]
        for excl in exclusions:
            ws.cell(row=current_row, column=1, value=excl).font = font_regular
            current_row += 1

        current_row += 2

        # 9. Materials Information Table
        ws.cell(row=current_row, column=1, value="Materials Information:").font = font_bold
        current_row += 1

        mat_hdr_row = current_row
        ws.cell(row=mat_hdr_row, column=1, value="Specification").font = font_header
        ws.cell(row=mat_hdr_row, column=1).fill = fill_floor
        ws.cell(row=mat_hdr_row, column=7, value="Total Qty").font = font_header
        ws.cell(row=mat_hdr_row, column=7).fill = fill_floor
        ws.cell(row=mat_hdr_row, column=8, value="Notes / Grout Specs").font = font_header
        ws.cell(row=mat_hdr_row, column=8).fill = fill_floor
        current_row += 1

        for idx, (sym, data) in enumerate(sorted(sym_totals.items()), 1):
            spec = project.material_specs.get(sym)
            desc_text = f"{idx}. {sym}: {spec.description}" if spec and spec.description else f"{idx}. {sym}: Standard Specification"
            
            row_idx = current_row
            ws.cell(row=row_idx, column=1, value=desc_text).font = font_regular
            
            # Linked formula or sum
            rows_list = item_rows_map.get(sym, [])
            if rows_list:
                formula_qty = "=" + "+".join([f"J{r}" for r in rows_list])
            else:
                formula_qty = data["qty"]
            
            q_c = ws.cell(row=row_idx, column=7, value=formula_qty)
            q_c.alignment = align_right
            q_c.number_format = '#,##0.00'
            q_c.font = font_bold

            note_text = spec.notes if spec and spec.notes else ""
            ws.cell(row=row_idx, column=8, value=note_text).font = font_sub
            current_row += 1

        current_row += 2
        ws.cell(row=current_row, column=1, value="Respectfully submitted,").font = font_bold
        current_row += 1
        sub_title = f"{project.estimator_name} - {project.estimator_title}" if project.estimator_name else "Estimating Department"
        ws.cell(row=current_row, column=1, value=sub_title).font = font_bold
        if project.bidder_company:
            current_row += 1
            ws.cell(row=current_row, column=1, value=project.bidder_company).font = font_bold
            if project.bidder_phone or project.bidder_email:
                current_row += 1
                contact_line = f"Tel: {project.bidder_phone} | Email: {project.bidder_email}" if project.bidder_phone and project.bidder_email else (project.bidder_phone or project.bidder_email)
                ws.cell(row=current_row, column=1, value=contact_line).font = font_sub

        wb.save(output_path)
        return output_path

    @staticmethod
    def get_dynamic_sow_rows(project: ProjectTakeoff) -> list:
        trade_groups: Dict[str, Dict[str, Dict[Any, Any]]] = {}
        for room in project.rooms:
            for item in room.items:
                trade_name = item.trade or "Tile & Stone"
                if trade_name not in trade_groups:
                    trade_groups[trade_name] = {}
                
                is_common = any(k in room.room_name.upper() or k in room.floor_name.upper() for k in ["LOBBY", "TERRACE", "ROOF", "MAIL", "CORRIDOR", "1ST FL"])
                area_name = "Common Areas & Building Perimeter" if is_common else "Apartment Units (Residential Scope)"
                if len(project.rooms) <= 8:
                    area_name = f"{room.floor_name} - {room.room_name}"
                    
                if area_name not in trade_groups[trade_name]:
                    trade_groups[trade_name][area_name] = {}
                    
                key = (item.symbol, item.finish_type, item.material_type, item.unit)
                if key not in trade_groups[trade_name][area_name]:
                    spec = project.material_specs.get(item.symbol)
                    notes = spec.notes if (spec and spec.notes) else (spec.description if spec else "")
                    trade_groups[trade_name][area_name][key] = {
                        "symbol": item.symbol,
                        "finish_type": item.finish_type,
                        "material_type": item.material_type,
                        "unit": item.unit,
                        "qty": 0.0,
                        "unit_price": item.material_price + item.labor_price,
                        "notes": notes
                    }
                trade_groups[trade_name][area_name][key]["qty"] += item.quantity

        sow_rows = []
        for trade_name, areas in trade_groups.items():
            sow_rows.append(("SEC", f"{trade_name.upper()} SCOPE", 0.0, "", ""))
            for area_name, items in areas.items():
                sow_rows.append(("SUB", area_name, 0.0, "", ""))
                for key, it in items.items():
                    desc = f"{it['finish_type']} ({it['symbol']})" if it['symbol'] else it['finish_type']
                    total_cost = it['qty'] * it['unit_price']
                    qty_str = f"{it['qty']:,.2f} {it['unit']}" if it['qty'] else it['unit']
                    sow_rows.append(("ITEM", desc, total_cost, qty_str, it['notes']))
                    
        return sow_rows

    @staticmethod
    def generate_sow_excel(project: ProjectTakeoff, output_path: str) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bid SOW Matrix"
        ExcelProposalGenerator._populate_sow_sheet(ws, project)
        wb.save(output_path)
        return output_path

    @staticmethod
    def _populate_sow_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, project: ProjectTakeoff):
        # Page Setup: Portrait & Fit to 1 Page Wide
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2

        ws.print_title_rows = '5:5'

        # Column widths (compact for portrait printing)
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 28

        # Fonts & Styles
        font_title = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
        font_sec_hdr = Font(name="Calibri", size=11, bold=True, color="0F172A")
        font_sub_hdr = Font(name="Calibri", size=10, bold=True, color="1E293B")
        font_item = Font(name="Calibri", size=10, color="334155")
        font_total = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
        font_notes = Font(name="Calibri", size=9, italic=True, color="64748B")

        fill_sec = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        fill_sub = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        fill_price = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        fill_grand = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        border_thin = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        # Title Block
        ws['A1'] = "PROJECT NAME"
        ws['B1'] = project.project_name
        ws['A1'].font = font_sec_hdr
        ws['B1'].font = font_title

        ws['A2'] = "BIDDER / CONTRACTOR"
        bidder_display = f"{project.bidder_company} ({project.estimator_name})" if project.bidder_company and project.estimator_name else (project.bidder_company or project.estimator_name or "Commercial Subcontractor")
        ws['B2'] = bidder_display
        ws['A2'].font = font_sec_hdr
        ws['B2'].font = font_item

        ws['A3'] = "DATE / CONTACT"
        contact_display = f"{project.date_str or datetime.date.today().strftime('%m/%d/%Y')} | Tel: {project.bidder_phone or 'N/A'} | Email: {project.bidder_email or 'N/A'}"
        ws['B3'] = contact_display
        ws['A3'].font = font_sec_hdr
        ws['B3'].font = font_item

        # Headers
        headers = [('A5', 'ITEM / SCOPE OF WORK'), ('B5', 'TOTAL ($)'), ('C5', 'QUANTITY / UNIT'), ('D5', 'SCOPE & DRAWING SPEC NOTES')]
        for cell_ref, text in headers:
            c = ws[cell_ref]
            c.value = text
            c.font = font_sec_hdr
            c.fill = fill_sec
            c.alignment = align_center if cell_ref != 'A5' else align_left
            c.border = border_thin

        sow_rows = ExcelProposalGenerator.get_dynamic_sow_rows(project)

        curr_r = 6
        item_rows = []

        for rtype, name, val, cost_sf, note in sow_rows:
            c_a = ws.cell(row=curr_r, column=1, value=name)
            c_b = ws.cell(row=curr_r, column=2)
            c_c = ws.cell(row=curr_r, column=3)
            c_d = ws.cell(row=curr_r, column=4, value=note)

            c_a.border = border_thin
            c_b.border = border_thin
            c_c.border = border_thin
            c_d.border = border_thin
            c_d.font = font_notes

            if rtype == "SEC":
                c_a.font = font_sec_hdr
                c_a.fill = fill_sec
                c_b.fill = fill_sec
                c_c.fill = fill_sec
                c_d.fill = fill_sec
            elif rtype == "SUB":
                c_a.font = font_sub_hdr
                c_a.fill = fill_sub
                c_b.fill = fill_sub
                c_c.fill = fill_sub
                c_d.fill = fill_sub
            else:
                c_a.font = font_item
                c_b.value = val
                c_b.font = font_item
                c_b.fill = fill_price
                c_b.alignment = align_right
                c_b.number_format = '$#,##0.00'

                c_c.value = cost_sf
                c_c.font = font_item
                c_c.alignment = align_right

                item_rows.append(curr_r)

            curr_r += 1

        curr_r += 1

        # Summary Rows
        subtotal_row = curr_r
        ws.cell(row=subtotal_row, column=1, value="SUBTOTAL").font = font_total
        ws.cell(row=subtotal_row, column=1).fill = fill_sub
        if item_rows:
            c_sub = ws.cell(row=subtotal_row, column=2, value="=" + "+".join([f"B{r}" for r in item_rows]))
        else:
            c_sub = ws.cell(row=subtotal_row, column=2, value=0.0)
        c_sub.font = font_total
        c_sub.fill = fill_sub
        c_sub.alignment = align_right
        c_sub.number_format = '$#,##0.00'
        for col in range(1, 5):
            ws.cell(row=subtotal_row, column=col).border = border_thin
        curr_r += 1

        # Overhead & Profit (10%)
        oh_row = curr_r
        ws.cell(row=oh_row, column=1, value="Overhead & Profit (10%)").font = font_item
        c_oh = ws.cell(row=oh_row, column=2, value=f"=B{subtotal_row}*0.10")
        c_oh.font = font_item
        c_oh.alignment = align_right
        c_oh.number_format = '$#,##0.00'
        for col in range(1, 5):
            ws.cell(row=oh_row, column=col).border = border_thin
        curr_r += 1

        # Insurance (3%)
        ins_row = curr_r
        ws.cell(row=ins_row, column=1, value="Insurance (3%)").font = font_item
        c_ins = ws.cell(row=ins_row, column=2, value=f"=B{subtotal_row}*0.03")
        c_ins.font = font_item
        c_ins.alignment = align_right
        c_ins.number_format = '$#,##0.00'
        for col in range(1, 5):
            ws.cell(row=ins_row, column=col).border = border_thin
        curr_r += 1

        # Grand Total
        gt_row = curr_r
        ws.cell(row=gt_row, column=1, value="GRAND TOTAL").font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
        ws.cell(row=gt_row, column=1).fill = fill_grand
        c_gt = ws.cell(row=gt_row, column=2, value=f"=B{subtotal_row}+B{oh_row}+B{ins_row}")
        c_gt.font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
        c_gt.fill = fill_grand
        c_gt.alignment = align_right
        c_gt.number_format = '$#,##0.00'
        for col in range(1, 5):
            ws.cell(row=gt_row, column=col).border = border_thin
