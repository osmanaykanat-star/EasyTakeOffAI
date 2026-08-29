import os
import openpyxl
from backend.engines.sample_data import get_zeta_sample_project
from backend.engines.excel_generator import ExcelProposalGenerator

def test_excel_generation():
    project = get_zeta_sample_project()
    out_path = os.path.join(os.path.dirname(__file__), "test_output.xlsx")
    ExcelProposalGenerator.generate_excel(project, out_path)
    
    assert os.path.exists(out_path), "Excel output file was not created"
    
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    
    assert ws['A1'].value == "PROPOSAL"
    assert "Zeta Charter Schools" in ws['B8'].value
    assert ws['A20'].value == "BASE BID"
    print("test_excel_generation passed successfully!")

if __name__ == "__main__":
    test_excel_generation()
